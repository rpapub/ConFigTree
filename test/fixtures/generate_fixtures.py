"""
Generate test fixture .xlsx files for ConFormMold.
Run with: uv run generate_fixtures.py
"""

import datetime
import pathlib
import openpyxl
from openpyxl.styles import Font

OUTPUT_DIR = pathlib.Path(__file__).parent

HEADER_STANDARD = ["Name", "Value", "Description"]
HEADER_ASSET    = ["Name", "Asset", "OrchestratorAssetFolder", "Description"]


def write_sheet(ws, header, rows):
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)


# ---------------------------------------------------------------------------
# Config_Basic.xlsx — strings and ints only, no `using System;` expected
# ---------------------------------------------------------------------------
def make_basic():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["OrchestratorQueueName", "BasicQueue",      "Orchestrator queue name."],
        ["OrchestratorFolderPath", "RPA/Basic",      "Orchestrator folder path."],
        ["MaxItemsPerRun",         10,                "Max items to process per run."],
        ["RetryCount",             3,                 "Number of retries on failure."],
        ["LogPrefix",             "BASIC",            "Prefix for log messages."],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["MaxRetryNumber",              0,     "Must be 0 when using Orchestrator queues."],
        ["MaxConsecutiveSystemExceptions", 3,  "Stop job after this many consecutive errors."],
        ["RetryNumberGetTransactionItem",  2,  "Retries for GetTransactionItem activity."],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [])  # empty — empty class expected

    wb.save(OUTPUT_DIR / "Config_Basic.xlsx")
    print("Created Config_Basic.xlsx")


# ---------------------------------------------------------------------------
# Config_Types.xlsx — all supported C# types, `using System;` expected
# ---------------------------------------------------------------------------
def make_types():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["FeatureName",       "TypesDemo",                          "string"],
        ["MaxItems",          42,                                   "int"],
        ["Threshold",         3.14,                                 "double"],
        ["IsEnabled",         True,                                 "bool"],
        ["CutoffDate",        datetime.date(2025, 12, 31),          "DateOnly — date only, time is 00:00:00"],
        ["ScheduledAt",       datetime.datetime(2025, 6, 15, 9, 30),"DateTime — has time component"],
        ["DailyRunTime",      datetime.time(8, 0, 0),               "TimeOnly — time only, no date"],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["Pi",                3.14159,                              "double — mathematical constant"],
        ["MaxRetryNumber",    0,                                    "int"],
        ["StrictMode",        False,                                "bool"],
        ["ExpiresOn",         datetime.date(2026, 1, 1),            "DateOnly"],
        ["CreatedAt",         datetime.datetime(2024, 3, 1, 12, 0), "DateTime"],
        ["WindowOpen",        datetime.time(9, 0, 0),               "TimeOnly"],
        ["WindowClose",       datetime.time(17, 30, 0),             "TimeOnly"],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [
        ["CredentialM365",    "cred_m365_types",  "Shared",  "M365 service credential."],
        ["CredentialFtp",     "cred_ftp_types",   "Shared",  "FTP server credential."],
    ])

    wb.save(OUTPUT_DIR / "Config_Types.xlsx")
    print("Created Config_Types.xlsx")


# ---------------------------------------------------------------------------
# Config_Assets.xlsx — focus on asset schema and OrchestratorAsset output
# ---------------------------------------------------------------------------
def make_assets():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["Environment",  "UAT",   "Deployment environment."],
        ["LogLevel",     "Info",  "Minimum log level."],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["MaxRetryNumber", 0, "Must be 0 when using Orchestrator queues."],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [
        ["CredentialSap",     "cred_sap_uat",       "SAP",     "SAP system credential."],
        ["CredentialM365",    "cred_m365_uat",      "Shared",  "M365 credential."],
        ["CredentialFtp",     "cred_ftp_uat",       "FTP",     "FTP server credential."],
        ["ApiKeyPayment",     "apikey_payment_uat", "API",     "Payment gateway API key."],
    ])

    wb.save(OUTPUT_DIR / "Config_Assets.xlsx")
    print("Created Config_Assets.xlsx")


if __name__ == "__main__":
    make_basic()
    make_types()
    make_assets()
    print("Done.")
